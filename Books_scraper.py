import requests
from bs4 import BeautifulSoup
import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

all_books = []
for page in range(1, 3):
    url = f"https://books.toscrape.com/catalogue/page-{page}.html"
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'html.parser')
    for book in soup.find_all('article', class_='product_pod'):
        title = book.h3.a['title']
        price = book.find('p', class_='price_color').text
        stock = book.find('p', class_='instock availability').text.strip()
        all_books.append({'title': title, 'price': price, 'stock': stock})
df = pd.DataFrame(all_books, columns=['title', 'price', 'stock'])
folder_name = "books_data"
os.makedirs(folder_name, exist_ok=True)
today = datetime.now().strftime("%Y-%m-%d")
file_path = f"{folder_name}/books_{today}.xlsx"
df.to_excel(file_path, index=False)
wb = load_workbook(file_path)
ws = wb.active
for col in range(1, ws.max_column + 1):
    ws.cell(row=1, column=col).font = Font(bold=True)
    ws.column_dimensions[get_column_letter(col)].width = 30

wb.save(file_path)
print(f"Report saved: {file_path}")
